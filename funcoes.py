import random 
import json
from openpyxl import workbook as opyxl

def convert_txt_list(filename:str)->list:
    """
    transforme txt file with medicine names to a list

    param filename: the name of the .txt file that need to be converted
    return: list with txt information
    """
    with open(filename, 'r', encoding='utf-8') as f:
        return [line.strip() for line in f.readlines() if line.strip()]

def generate_intercection_matriz(medicines:list)->list:
    """
    creat a list to generate random reaction between the medicines using numbers 0 to 6

    param medicine: list with all medicines that will react with each other
    return: matriz whit all reactions
    """
    size = len(medicines)
    matriz = [[0]*size for _ in range(size)]

    for i in range(size):
        for j in range(i+1, size):
            valor = random.randint(0, 6)
            matriz[i][j] = valor 
            matriz[j][i] = valor

    return matriz

def gerar_excel(filename:str, medicines:list, matriz:list)->None:
    """
    generate excel to represente the reactions created in "generate_intercection_matriz"

    param filename: the name of the resulting .xlsx file
    param medicines: same list with name of medicines used in "generate_intercection_matriz"
    param matriz: matriz resulted of the function "generate_intercection_matriz"
    retrun: nothing
    """
    wb = opyxl.Workbook()
    ws = wb.active
    ws.title = "Interações"

    # escrever cabeçalhos das colunas
    for col, med in enumerate(medicines, start=2):
        ws.cell(row=1, column=col, value=med)

    # escrever cabeçalhos das linhas + valores
    for row, med in enumerate(medicines, start=2):
        ws.cell(row=row, column=1, value=med)
        for col in range(len(medicines)):
            ws.cell(row=row, column=col+2, value=matriz[row-2][col])

    wb.save(filename)
    return None

def creat_prescription(medicines:list)->dict:
    """
    randomly chose 1 to 10 people and for every one create a create a prescrito between 3 and 5 medicines

    param medicine: list of medicines
    retrun: list with medicine prescripted
    """
    people={ }
    
    for _ in range(random.randint(1,10)):
        prescription = []
        for _ in range(random.randint(3,5)):
            prescription.append(random.choice(medicines).strip())
        people[random.randint(100000000,999999999)] = prescription
    return people




def danger_avaliation(prescription:dict,matriz:list,medicines:list )->list:
    """
    check the highs risk reaction between the medicine in the prescription

    param medicine: list of medicines choosed
    param matriz: list with the level of danger of the reaction between de medicines
    retrun: value that represente the level of danger
    """
    max_list = []
    for key in prescription:
        danger = []
        max_danger= 0
        
        for i in range(len(prescription[key])):
            for j in range(i + 1,len(prescription[key])):                
                danger.append(matriz[medicines.index(prescription[key][i])][medicines.index(prescription[key][j])])
        for i in range(len(danger)):
            if max_danger < danger[i]:
                max_danger = danger[i]
        max_list.append(max_danger)

    
    return max_list

def danger_count(prescription:list,matriz:list)->int:
    """
    count the number of reactions between the medicines in the prescription

    param medicine: list of medicines choosed
    param matriz: list with the level of danger of the reaction between de medicines
    retrun: value that represente the number of reactions
    """
    count = 0
    for i in range(len(prescription)):
        for j in range(i + 1,len(prescription)):
            if matriz[i][j] != 0:
                count += 1
    return count

def  edit_json(dict:dict, filename:str)->None:
    """
    save the prescription in a .json file

    param dict: the prescription that will be saved
    param filename: the name of the resulting .json file
    retrun: nothing
    """
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(dict, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print(f"Error saving JSON: {e}")
    return None

def encription_utent_number(prescription:dict, key:int)->dict:
    """
    encript with cesar cipher the utent number in the prescription

    param dict: the prescription that will be encripted
    retrun: the prescription with the utent number encripted
    """
    for _ in range(len(prescription)):
        numero_utentes = []
        for i in prescription:            numero_utentes.append([int(d) for d in str(i)]) #transform the utent number in a list of digits
    ascii_listas = [[int(ord(str(digito))) + key % 256 for digito in sublista] for sublista in numero_utentes]  #transform the digits in ascii values and apply the caesar cipher
    no_ascii= [[chr(ascii_val) for ascii_val in sublista] for sublista in ascii_listas] #transform the ascii values back to characters
    lista_encriptada = [''.join(sublista) for sublista in no_ascii] #join the characters to form the encripted utent number
    encripted_dict = {lista_encriptada[i] : prescription[key] for i, key in enumerate(prescription)} 
    return encripted_dict


    
def description_utent_number(prescription:dict, chave:int)->dict:
    """
    descript with cesar cipher the utent number in the prescription

    param dict: the prescription that will be descripted
    retrun: the prescription with the utent number descripted
    """
    for _ in range(len(prescription)):
        numero_utentes = []
        for i in prescription:
            numero_utentes.append(list(str(i))) #transform the encripted utent number in a list of characters
    ascii_listas = [[int(ord(str(digito))) - chave % 256 for digito in sublista] for sublista in numero_utentes] #transform the characters in ascii values and apply the caesar cipher in reverse
    no_ascii= [[chr(ascii_val) for ascii_val in sublista] for sublista in ascii_listas] #transform the ascii values back to characters
    lista_descriptada = [''.join(sublista) for sublista in no_ascii] #join the characters to form the descripted utent number
    descripted_dict = {lista_descriptada[i] : prescription[key] for i, key in enumerate(prescription)}
    return descripted_dict 

def hashing_folding(prescription:dict, table_size:int, keyedhash:int)->list:
    """
    apply folding method to hash the utent number in the prescription

    param dict: the prescription that will be hashed
    param table_size: the size of the hash table
    retrun: the prescription with the utent number hashed
    """
    hashed_dict = []
    for key in prescription:
        key_str = str(key)
        fold_sum = sum(int(ord(digit)) for digit in key_str) + keyedhash  # Sum of digits
        hashed_key = fold_sum % table_size  # Modulo by table size
        hashed_dict.append(hashed_key)
    return hashed_dict

medicines = convert_txt_list("medicines.txt")
matriz = generate_intercection_matriz(medicines)
gerar_excel("interacoes.xlsx", medicines, matriz)
prescription = creat_prescription(medicines)
print(danger_avaliation(prescription,matriz,medicines))
encripted_prescription = encription_utent_number(prescription, 3)
print(encripted_prescription)
edit_json(encripted_prescription, "prescription.json")
print(description_utent_number(encripted_prescription, 3))
print([hex(hexadecimal) for hexadecimal in hashing_folding((prescription), 256, 123987)])